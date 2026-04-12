"""
Data loader module for handling large Excel files efficiently.
Supports chunked reading, data preprocessing, and relationship detection.
"""

import pandas as pd
import openpyxl
from typing import List, Dict, Any, Optional, Tuple
import logging
from pathlib import Path

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelDataLoader:
    """Efficient loader for large Excel files with chunking support."""

    def __init__(self, file_paths: List[str], chunk_size: int = 10000):
        """
        Initialize the data loader.

        Args:
            file_paths: List of paths to Excel files
            chunk_size: Number of rows to process at once
        """
        self.file_paths = file_paths
        self.chunk_size = chunk_size
        self.dataframes = {}
        self.metadata = {}

    def analyze_structure(self) -> Dict[str, Any]:
        """
        Analyze the structure of Excel files without loading all data.

        Returns:
            Dictionary containing metadata about each file and sheet
        """
        analysis = {}

        for file_path in self.file_paths:
            logger.info(f"Analyzing: {file_path}")
            file_name = Path(file_path).name
            analysis[file_name] = {
                'sheets': {},
                'total_rows': 0,
                'relationships': []
            }

            # Get sheet names
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names

            for sheet_name in sheet_names:
                # Read only first few rows to get schema
                df_sample = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)

                # Get full row count efficiently
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb[sheet_name]
                row_count = ws.max_row - 1  # Subtract header row
                col_count = ws.max_column
                wb.close()

                analysis[file_name]['sheets'][sheet_name] = {
                    'columns': list(df_sample.columns),
                    'dtypes': {col: str(dtype) for col, dtype in df_sample.dtypes.items()},
                    'row_count': row_count,
                    'col_count': col_count,
                    'sample_data': df_sample.head(3).to_dict('records')
                }

                analysis[file_name]['total_rows'] += row_count

            logger.info(f"File: {file_name}, Total rows: {analysis[file_name]['total_rows']}")

        self.metadata = analysis
        return analysis

    def detect_relationships(self) -> List[Dict[str, Any]]:
        """
        Detect potential relationships between tables based on common columns.

        Returns:
            List of potential relationships
        """
        relationships = []

        if not self.metadata:
            self.analyze_structure()

        # Compare columns across all sheets
        all_sheets = []
        for file_name, file_data in self.metadata.items():
            for sheet_name, sheet_data in file_data['sheets'].items():
                all_sheets.append({
                    'file': file_name,
                    'sheet': sheet_name,
                    'columns': set(sheet_data['columns'])
                })

        # Find common columns (potential foreign keys)
        for i, sheet1 in enumerate(all_sheets):
            for sheet2 in all_sheets[i+1:]:
                common_cols = sheet1['columns'].intersection(sheet2['columns'])
                if common_cols:
                    relationships.append({
                        'table1': f"{sheet1['file']}:{sheet1['sheet']}",
                        'table2': f"{sheet2['file']}:{sheet2['sheet']}",
                        'common_columns': list(common_cols),
                        'relationship_type': 'potential_join'
                    })

        return relationships

    def load_data_chunked(self, file_path: str, sheet_name: str) -> pd.DataFrame:
        """
        Load data from Excel file in chunks to handle large datasets.

        Args:
            file_path: Path to Excel file
            sheet_name: Name of the sheet to load

        Returns:
            Complete DataFrame
        """
        logger.info(f"Loading {file_path}:{sheet_name} in chunks")

        chunks = []

        # Read in chunks
        excel_file = pd.ExcelFile(file_path)

        # Get total rows
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[sheet_name]
        total_rows = ws.max_row - 1
        wb.close()

        # Read in chunks
        for i in range(0, total_rows, self.chunk_size):
            chunk = pd.read_excel(
                excel_file,
                sheet_name=sheet_name,
                skiprows=range(1, i + 1) if i > 0 else None,
                nrows=self.chunk_size
            )
            chunks.append(chunk)
            logger.info(f"Loaded chunk {len(chunks)}: {len(chunk)} rows")

        # Combine all chunks
        df = pd.concat(chunks, ignore_index=True)
        logger.info(f"Total rows loaded: {len(df)}")

        return df

    def load_all_data(self) -> Dict[str, Dict[str, pd.DataFrame]]:
        """
        Load all data from all files and sheets.

        Returns:
            Nested dictionary: {file_name: {sheet_name: dataframe}}
        """
        all_data = {}

        for file_path in self.file_paths:
            file_name = Path(file_path).name
            all_data[file_name] = {}

            excel_file = pd.ExcelFile(file_path)
            for sheet_name in excel_file.sheet_names:
                df = self.load_data_chunked(file_path, sheet_name)
                all_data[file_name][sheet_name] = df

        self.dataframes = all_data
        return all_data

    def prepare_documents_for_embedding(self, max_rows_per_chunk: int = 50) -> List[Dict[str, Any]]:
        """
        Prepare data for embedding generation.
        Converts rows into text documents suitable for vector storage.

        Args:
            max_rows_per_chunk: Number of rows to combine into one document

        Returns:
            List of documents with metadata
        """
        documents = []

        if not self.dataframes:
            self.load_all_data()

        for file_name, sheets in self.dataframes.items():
            for sheet_name, df in sheets.items():
                # Process in chunks
                for chunk_idx in range(0, len(df), max_rows_per_chunk):
                    chunk_df = df.iloc[chunk_idx:chunk_idx + max_rows_per_chunk]

                    # Convert chunk to text representation
                    text_parts = []
                    text_parts.append(f"File: {file_name}, Sheet: {sheet_name}")
                    text_parts.append(f"Columns: {', '.join(str(col) for col in chunk_df.columns)}")
                    text_parts.append("Data:")

                    # Create a summary of the chunk
                    for idx, row in chunk_df.iterrows():
                        row_text = ", ".join([f"{str(col)}: {str(val)}" for col, val in row.items() if pd.notna(val)])
                        text_parts.append(row_text)

                    document = {
                        'text': "\n".join(text_parts),
                        'metadata': {
                            'file': file_name,
                            'sheet': sheet_name,
                            'start_row': chunk_idx,
                            'end_row': min(chunk_idx + max_rows_per_chunk, len(df)),
                            'num_rows': len(chunk_df)
                        },
                        'data': chunk_df.to_dict('records')
                    }

                    documents.append(document)

        logger.info(f"Prepared {len(documents)} documents for embedding")
        return documents

    def get_summary_statistics(self) -> Dict[str, Any]:
        """
        Generate summary statistics for all loaded data.

        Returns:
            Dictionary of statistics
        """
        if not self.dataframes:
            self.load_all_data()

        stats = {}

        for file_name, sheets in self.dataframes.items():
            stats[file_name] = {}
            for sheet_name, df in sheets.items():
                stats[file_name][sheet_name] = {
                    'row_count': len(df),
                    'column_count': len(df.columns),
                    'columns': list(df.columns),
                    'numeric_columns': list(df.select_dtypes(include=['number']).columns),
                    'categorical_columns': list(df.select_dtypes(include=['object']).columns),
                    'missing_values': df.isnull().sum().to_dict(),
                    'memory_usage_mb': df.memory_usage(deep=True).sum() / 1024 / 1024
                }

                # Add basic statistics for numeric columns
                numeric_stats = df.describe().to_dict()
                stats[file_name][sheet_name]['numeric_stats'] = numeric_stats

        return stats


if __name__ == "__main__":
    # Example usage
    loader = ExcelDataLoader([
        "actuarial_life_data_file1.xlsx",
        "actuarial_life_data_file2.xlsx"
    ])

    # Analyze structure
    print("="*60)
    print("ANALYZING EXCEL FILE STRUCTURE")
    print("="*60)
    analysis = loader.analyze_structure()

    import json
    print(json.dumps(analysis, indent=2, default=str))

    # Detect relationships
    print("\n" + "="*60)
    print("DETECTING RELATIONSHIPS")
    print("="*60)
    relationships = loader.detect_relationships()
    print(json.dumps(relationships, indent=2))

    # Get summary statistics
    print("\n" + "="*60)
    print("SUMMARY STATISTICS")
    print("="*60)
    stats = loader.get_summary_statistics()
    print(json.dumps(stats, indent=2, default=str))
