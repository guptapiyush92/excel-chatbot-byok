import openpyxl
import sys

def analyze_excel(filename):
    print(f"\n{'='*60}")
    print(f"Analyzing: {filename}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
    print(f"\nSheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")

        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"Dimensions: {max_row} rows x {max_col} columns")

        # Get headers (first row)
        headers = []
        for col in range(1, min(max_col + 1, 20)):  # Limit to first 20 columns
            cell_value = ws.cell(1, col).value
            headers.append(cell_value)
        print(f"Headers: {headers}")

        # Get first 5 data rows
        print("\nFirst 5 rows of data:")
        for row in range(2, min(7, max_row + 1)):
            row_data = []
            for col in range(1, min(max_col + 1, 10)):
                cell_value = ws.cell(row, col).value
                row_data.append(cell_value)
            print(f"Row {row}: {row_data}")

    wb.close()

if __name__ == "__main__":
    analyze_excel("actuarial_life_data_file1.xlsx")
    analyze_excel("actuarial_life_data_file2.xlsx")
